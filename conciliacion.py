"""
CONCILIACIÓN BANCARIA ADA4
===========================
Concilia automáticamente los movimientos del extracto bancario con:
  - Los gastos cargados en el sistema ADA4 (InformeGastos_*.xls)
  - Los cobros cargados en el sistema ADA4 (CobranzaRapida_*.xlsx)

LÓGICA DE MATCHING:
  COBROS: 5 estrategias en orden de prioridad:
    1. QR: matching por importe bruto (col H del extracto).
    2. Match único: importe del banco == movimiento ADA de una UF.
    3. Cuotas: 2-3 cobros cuya suma coincide con movimiento/deuda de la UF.
       (3a: mismo CUIT; 3b: sin CUIT + centavos exactos; 3c: suma exacta + anchor centavo)
    4. Centavos: los centavos del importe identifican la UF (truco local).
    5. Sin match.

  GASTOS:
    - Match por importe con tolerancia de centavos (0.50 por redondeos)
    - Si hay ambigüedad (varios gastos con mismo importe), no aparece en match

SALIDA:
  1. Reporte de conciliación (ConciliacionBancaria_YYYY-MM-DD_HH-MM.xlsx)
  2. Extracto original con colores:
     - Verde: conciliado
     - Amarillo: diferencia de importe o ambiguo
     - Rojo: sin match

CONFIGURACIÓN: Editá las rutas abajo.
"""

import subprocess, re, sys, shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
#  CONFIGURACIÓN - EDITÁ ESTAS RUTAS
# ============================================================

EXTRACTO_BANCO   = r"extracto.xlsx"

# Carpeta donde están los archivos descargados del sistema ADA4
# El script busca automáticamente el archivo más reciente que empiece con
# "InformeGastos" y "CobranzaRapida" en esa carpeta.
CARPETA_ADA4     = r"."

CARPETA_SALIDA   = r"."   # donde se guardan el reporte y el extracto marcado

# Tolerancia para matching de importes (en pesos)
TOLERANCIA_GASTO = 0.50   # centavos de redondeo en liquidaciones
TOLERANCIA_COBRO = 0.02   # los cobros deben ser muy exactos

# ============================================================

FILL_VERDE    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_NARANJA  = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
FILL_ROJO     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER   = Font(bold=True, color="FFFFFF")


def encontrar_libreoffice():
    for p in [r"C:\Program Files\LibreOffice\program\soffice.exe",
              r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"]:
        if Path(p).exists():
            return p
    return None


def buscar_archivo_por_prefijo(carpeta, prefijo, extensiones):
    """
    Busca en `carpeta` el archivo más reciente que empiece con `prefijo`
    y termine con alguna de las `extensiones`. Devuelve la ruta completa
    o None si no encuentra nada.
    """
    carpeta = Path(carpeta)
    if not carpeta.exists():
        return None
    candidatos = []
    for archivo in carpeta.iterdir():
        if not archivo.is_file():
            continue
        nombre = archivo.name
        if not nombre.startswith(prefijo):
            continue
        if not any(nombre.lower().endswith(ext.lower()) for ext in extensiones):
            continue
        candidatos.append(archivo)
    if not candidatos:
        return None
    # Devolver el más reciente por fecha de modificación
    candidatos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return str(candidatos[0])


def convertir_xls(ruta_xls):
    soffice = encontrar_libreoffice()
    if not soffice:
        raise FileNotFoundError(
            "LibreOffice no encontrado. Instalalo desde https://www.libreoffice.org/")
    xls = Path(ruta_xls)
    r = subprocess.run(
        [soffice, '--headless', '--convert-to', 'xlsx',
         str(xls), '--outdir', str(xls.parent)],
        capture_output=True, text=True)
    out = xls.parent / (xls.stem + '.xlsx')
    if not out.exists():
        raise RuntimeError(f"Falló la conversión:\n{r.stderr}")
    return str(out)


def leer_extracto(ruta):
    """Lee el extracto bancario. Devuelve lista de dicts con los movimientos.
    Para COBROS QR (descripción contiene 'CREDITO POR COBRO QR'), también
    captura el importe bruto de la columna H, que contiene el valor original
    pagado por el propietario (antes de deducir la comisión QR).
    """
    wb = load_workbook(ruta, read_only=True)
    ws = wb.active
    movimientos = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        fecha, comp, conc, desc, imp = row[0], row[1], row[2], row[3], row[4]
        if imp is None or isinstance(imp, str):
            continue
        descripcion = str(desc or '').strip()
        item = {
            'fila_banco':  i,
            'fecha':       fecha,
            'comprobante': str(comp or '').strip(),
            'concepto':    str(conc or '').strip(),
            'descripcion': descripcion,
            'importe':     float(imp),
            'tipo':        'cobro' if imp > 0 else 'gasto',
            'es_qr':       False,
            'bruto_qr':    None,
        }
        # Detectar COBRO QR y capturar el bruto
        # Formato 1: bruto en columna H (col índice 7), comisión implícita
        # Formato 2: comisión negativa en columna F (col índice 5), bruto = neto + |comisión|
        if 'CREDITO POR COBRO QR' in descripcion.upper():
            h_val = row[7] if len(row) > 7 else None
            f_val = row[5] if len(row) > 5 else None
            if h_val is not None and not isinstance(h_val, str):
                item['es_qr'] = True
                item['bruto_qr'] = float(h_val)
            elif f_val is not None and not isinstance(f_val, str) and float(f_val) < 0:
                item['es_qr'] = True
                item['bruto_qr'] = float(imp) + abs(float(f_val))
        movimientos.append(item)
    wb.close()
    return movimientos


def leer_gastos_ada(ruta):
    """Lee el informe de gastos ADA4."""
    wb = load_workbook(ruta, read_only=True)
    ws = wb.active
    gastos = []
    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        # Col A=periodo, B=descripcion, D=importe, F=proveedor,
        # G=categoria, H=fecha_fact, I=nro_fact, J=id
        imp = row[3]
        if imp is None or isinstance(imp, str):
            continue
        gastos.append({
            'periodo':    row[0],
            'descripcion': str(row[1] or '').strip(),
            'importe':    float(imp),
            'proveedor':  str(row[5] or '').strip(),
            'categoria':  str(row[6] or '').strip(),
            'fecha_fact': row[7],
            'nro_fact':   str(row[8] or '').strip(),
            'id_gasto':   str(row[9] or '').strip(),
            'usado':      False,
        })
    wb.close()
    return gastos


def leer_cobros_ada(ruta):
    """Lee el listado de cobros de cobranza rápida."""
    wb = load_workbook(ruta, read_only=True)
    ws = wb.active
    cobros = []
    for row in ws.iter_rows(values_only=True):
        uf = row[2] if len(row) > 2 else None
        # Saltear encabezados (UF contiene el texto "UF" o None)
        if uf is None or str(uf).strip() in ('', 'UF'):
            continue
        uf_str = str(uf).strip()
        if not any(c.isdigit() for c in uf_str):
            continue
        deuda_ant = row[6]
        if deuda_ant is None or isinstance(deuda_ant, str):
            continue
        cobros.append({
            'uf':            uf_str,
            'depto':         str(row[3] or '').strip(),
            'piso':          str(row[4] or '').strip(),
            'propietario':   str(row[5] or '').strip(),
            'deuda_anterior': float(deuda_ant),
            'movimiento':    float(row[7]) if row[7] is not None else 0,
            'usado':         False,
        })
    wb.close()
    return cobros


def extraer_cuit(descripcion):
    """Extrae el CUIT/CUIL de una descripción del banco si existe."""
    m = re.search(r'-?(\d{11})-?', descripcion)
    if m:
        return m.group(1)
    m = re.search(r'(\d{2}-\d{8}-\d{1})', descripcion)
    if m:
        return m.group(1).replace('-', '')
    return None


def extraer_nombre_pagador(descripcion):
    """
    Extrae el nombre/apellido del pagador de la descripción del banco.
    Ej: 'TRANSF.INTER.DISTINTO TIT.-20273444427-PORCEL/ANDRUS' → 'PORCEL'
         'CREDIN - Crédito Inmediato Distinto titular-20439355573-FELIPE REBASSIO' → 'REBASSIO'
    Devuelve el token más largo en mayúsculas después del CUIT (normalmente el apellido).
    """
    # Sacar el CUIT si existe
    desc = re.sub(r'\d{11}', '', descripcion).upper()
    # Quedarse con caracteres de letras y separadores típicos
    desc = re.sub(r'[/,.\-]', ' ', desc)
    # Tokens de 3+ letras
    tokens = [t for t in desc.split() if len(t) >= 3 and t.isalpha()]
    # Filtrar palabras comunes del banco
    excluir = {'TRANSF', 'INTER', 'DISTINTO', 'TIT', 'CREDIN', 'CREDITO',
               'INMEDIATO', 'TITULAR', 'COBRO', 'TRANSFERENCIA', 'DEBITO',
               'DIRECTO', 'PAGO', 'AUTOMATICO', 'SERVICIOS', 'INTERB',
               'PAGADO', 'HABER', 'DEBIT', 'CREDIT', 'DEL', 'LAS', 'LOS',
               'LAVALLE', 'PROVINCIA', 'BAK', 'SIRO', 'LINK', 'PAGOS'}
    tokens = [t for t in tokens if t not in excluir]
    # El apellido suele ser el primer token largo (3+ letras)
    return tokens[0] if tokens else None


def obtener_centavos(importe):
    """Devuelve los centavos como entero (457551.09 → 9)."""
    return round((importe - int(importe)) * 100)


def es_cobro_qr(descripcion):
    """Detecta si un cobro viene de Cobro QR (tiene comisión aplicada)."""
    return 'CREDITO POR COBRO QR' in (descripcion or '').upper()


def conciliar_cobros(cobros_banco, cobros_ada):
    """
    Matchea cada cobro del banco con una UF del ADA4.

    Filosofía: la columna "Movimientos" de ADA es la verdad oficial sobre cuánto
    recibió cada UF. El script intenta reconstruir qué cobros del banco
    corresponden al movimiento registrado en ADA.

    Estrategias (en orden de prioridad):
      1. COBRO QR: matching especial por importe bruto (col H del extracto).
      2. MATCH ÚNICO: el importe del banco coincide con el movimiento ADA de una UF.
         Se ejecuta antes que las cuotas para que cobros individuales exactos no sean
         consumidos por agrupaciones erróneas.
      3. PAGO EN CUOTAS: 2-3 cobros cuya suma coincide con movimiento/deuda de la UF.
         3a: mismo CUIT, con centavos de la suma apuntando a la UF.
         3b: cualquier CUIT, centavos de la suma apuntan a la UF, tolerancia estricta.
         3c: cualquier CUIT, suma exacta con movimiento ADA + al menos un anchor de centavos.
      4. CENTAVOS: truco de decimales por UF (.02 → UF 0002), cuando no tenemos
         confirmación en ADA.
      5. SIN MATCH: el cobro no tiene contraparte en ADA.
    """
    resultados = []
    procesados = set()   # índices de cobros ya resueltos

    # === Helper: importe "real" a matchear (bruto para QR, importe normal para resto) ===
    def imp_match(cb):
        if cb.get('es_qr') and cb.get('bruto_qr') is not None:
            return cb['bruto_qr']
        return cb['importe']

    # Tolerancia para sumas de cuotas (mayor que la individual por acumulación
    # de redondeos centavo a centavo)
    TOLERANCIA_CUOTAS = 1.00

    # === Estrategia 1: COBROS QR ===
    # Los QR matchean contra la deuda esperada (los centavos del bruto apuntan a la UF)
    for idx, cb in enumerate(cobros_banco):
        if idx in procesados:
            continue
        if not (cb.get('es_qr') and cb.get('bruto_qr') is not None):
            continue
        bruto    = cb['bruto_qr']
        centavos = obtener_centavos(bruto)
        for ca in cobros_ada:
            uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')
            if uf_num == centavos and not ca['usado']:
                # Confirmar que la deuda ADA coincide con el bruto
                if abs(ca['deuda_anterior'] - bruto) <= TOLERANCIA_COBRO:
                    comision = abs(cb['importe'] - bruto)
                    detalle = (f"QR: bruto {bruto:,.2f}, comisión {comision:,.2f}")
                    resultados.append({
                        'banco':   cb,
                        'ada':     ca,
                        'estado':  'CONCILIADO',
                        'detalle': detalle,
                        'cuit':    extraer_cuit(cb['descripcion']),
                    })
                    ca['usado'] = True
                    procesados.add(idx)
                    break

    # === Estrategia 2: MATCH ÚNICO por importe = movimiento ADA ===
    # Se ejecuta ANTES que las cuotas para que un cobro que ya coincide
    # exactamente con una UF (por movimiento ADA) no sea "consumido" erróneamente
    # por una agrupación de cuotas que incluya un cobro espurio (ej: transferencia
    # de $1 por error que al sumarse da un match falso con otra UF).
    # Tolerancia ampliada a TOLERANCIA_CUOTAS ($1.00) para capturar pagos redondos
    # donde el banco recibió $N y ADA registró $N+1 (u otros ajustes pequeños).
    for ca in cobros_ada:
        if ca['usado'] or ca['movimiento'] == 0:
            continue
        movimiento_abs = abs(ca['movimiento'])
        uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')

        # Primero: candidatos con centavos coincidentes (más confiable)
        # Segundo: candidatos sin centavos coincidentes (usa tolerancia igual)
        cand_con_centavos = []
        cand_sin_centavos = []
        for idx, cb in enumerate(cobros_banco):
            if idx in procesados:
                continue
            imp = imp_match(cb)
            if abs(imp - movimiento_abs) <= TOLERANCIA_CUOTAS:
                if obtener_centavos(imp) == uf_num:
                    cand_con_centavos.append(idx)
                else:
                    cand_sin_centavos.append(idx)

        # Preferir los que coinciden por centavos
        candidatos = cand_con_centavos if cand_con_centavos else cand_sin_centavos

        if len(candidatos) == 1:
            idx = candidatos[0]
            cb  = cobros_banco[idx]
            imp = imp_match(cb)
            # Determinar el mensaje de detalle según cómo coincide con la deuda
            diff_deuda = abs(ca['deuda_anterior'] - imp)
            if diff_deuda <= TOLERANCIA_COBRO:
                detalle = ''
            elif abs(ca['deuda_anterior'] - movimiento_abs) <= TOLERANCIA_COBRO:
                detalle = ''  # movimiento coincide con deuda, diff es por redondeo
            elif movimiento_abs < ca['deuda_anterior']:
                saldo = ca['deuda_anterior'] - movimiento_abs
                detalle = (f"Pago parcial: pagó {movimiento_abs:,.2f} de "
                           f"{ca['deuda_anterior']:,.2f} (quedan {saldo:,.2f})")
            else:
                extra = movimiento_abs - ca['deuda_anterior']
                detalle = (f"Pagó de más: {movimiento_abs:,.2f} sobre deuda "
                           f"{ca['deuda_anterior']:,.2f} (saldo a favor: {extra:,.2f})")
            resultados.append({
                'banco':   cb,
                'ada':     ca,
                'estado':  'CONCILIADO',
                'detalle': detalle,
                'cuit':    extraer_cuit(cb['descripcion']),
            })
            ca['usado'] = True
            procesados.add(idx)

    # === Estrategia 3: PAGOS EN CUOTAS ===
    # Detecta 2+ cobros del banco cuya SUMA coincide con la deuda o el movimiento
    # ADA de una UF. Se ejecuta DESPUÉS de los matches individuales exactos para
    # evitar que cobros espurios (ej: transferencias de $1 por error) participen
    # en agrupaciones erróneas.
    #
    # Sub-estrategias:
    #   3a. Cuotas del mismo CUIT, preferentemente con centavos de la suma
    #       apuntando a la UF (señal fuerte de intención del pagador).
    #   3b. Cuotas sin CUIT común, solo si los centavos de la suma apuntan
    #       claramente a una UF específica.

    def intentar_match_cuotas(indices, ca, ref):
        """Genera los resultados para un match de cuotas encontrado."""
        total = sum(cobros_banco[i]['importe'] for i in indices)
        n_cuotas = len(indices)
        ref_monto = abs(ca['movimiento']) if ref == 'movimiento ADA' else ca['deuda_anterior']
        # CUIT común (si todos comparten) o vacío
        cuits = {extraer_cuit(cobros_banco[i]['descripcion']) for i in indices}
        cuits.discard(None)
        cuit_comun = cuits.pop() if len(cuits) == 1 else None
        descr_cuit = f"CUIT {cuit_comun}, " if cuit_comun else "distintos CUITs, "
        for i, idx in enumerate(
                sorted(indices, key=lambda x: cobros_banco[x]['fecha'] or datetime.min), 1):
            cb = cobros_banco[idx]
            detalle = (f"Pago en {n_cuotas} cuotas ({descr_cuit}total {total:,.2f} ≈ "
                       f"{ref} {ref_monto:,.2f}): cuota {i}/{n_cuotas} = {cb['importe']:,.2f}")
            resultados.append({
                'banco':   cb,
                'ada':     ca,
                'estado':  'PAGO EN CUOTAS',
                'detalle': detalle,
                'cuit':    extraer_cuit(cb['descripcion']),
            })
            procesados.add(idx)
        ca['usado'] = True

    # --- Estrategia 3a: cuotas del mismo CUIT, match por centavos de la suma ---
    grupos_cuit = {}
    for idx, cb in enumerate(cobros_banco):
        if cb.get('es_qr'):
            continue
        cuit = extraer_cuit(cb['descripcion'])
        if cuit:
            grupos_cuit.setdefault(cuit, []).append(idx)

    for cuit, indices in grupos_cuit.items():
        if len(indices) < 2:
            continue
        total = sum(cobros_banco[i]['importe'] for i in indices)
        centavos_suma = obtener_centavos(total)
        # Prioridad 1: buscar UF con mismos centavos que la suma
        uf_match = None
        ref_match = None
        for ca in cobros_ada:
            if ca['usado']:
                continue
            uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')
            if uf_num != centavos_suma:
                continue
            mov_abs = abs(ca['movimiento'])
            if abs(total - ca['deuda_anterior']) <= TOLERANCIA_CUOTAS:
                uf_match = ca
                ref_match = 'deuda esperada'
                break
            if mov_abs > 0 and abs(total - mov_abs) <= TOLERANCIA_CUOTAS:
                uf_match = ca
                ref_match = 'movimiento ADA'
                break
        # Prioridad 2: fallback a match por movimiento ADA (sin requerir centavos)
        # pero SOLO cuando los centavos individuales de TODAS las cuotas coinciden
        # con la UF (caso clásico: .19 + .19 = .38 pero cada cuota es .19)
        if uf_match is None:
            for ca in cobros_ada:
                if ca['usado']:
                    continue
                uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')
                # Verificar que todas las cuotas tengan los centavos de la UF
                if not all(obtener_centavos(cobros_banco[i]['importe']) == uf_num
                           for i in indices):
                    continue
                mov_abs = abs(ca['movimiento'])
                if mov_abs > 0 and abs(total - mov_abs) <= TOLERANCIA_CUOTAS:
                    uf_match = ca
                    ref_match = 'movimiento ADA'
                    break
                if abs(total - ca['deuda_anterior']) <= TOLERANCIA_CUOTAS:
                    uf_match = ca
                    ref_match = 'deuda esperada'
                    break
        if uf_match:
            intentar_match_cuotas(indices, uf_match, ref_match)

    # --- Estrategia 3b: cuotas sin CUIT común, solo si los centavos de la suma
    # apuntan claramente a una UF específica Y la suma coincide EXACTAMENTE con
    # el movimiento ADA de esa UF (tolerancia estricta para evitar falsos positivos
    # al sumar cobros espurios como transferencias de $1 por error). ---
    from itertools import combinations
    for ca in cobros_ada:
        if ca['usado'] or ca['movimiento'] == 0:
            continue
        uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')
        mov_abs = abs(ca['movimiento'])
        # Candidatos: cobros no procesados, no QR
        disponibles = [idx for idx in range(len(cobros_banco))
                       if idx not in procesados and not cobros_banco[idx].get('es_qr')]
        if len(disponibles) < 2:
            continue
        # Probar combinaciones de 2 y 3 cobros
        encontrado = False
        for n in (2, 3):
            if encontrado:
                break
            if len(disponibles) < n:
                continue
            for combo in combinations(disponibles, n):
                total = sum(cobros_banco[i]['importe'] for i in combo)
                # Exigir que los centavos de la suma apunten a esta UF
                if obtener_centavos(total) != uf_num:
                    continue
                # Exigir coincidencia EXACTA con movimiento ADA (tolerancia estricta)
                # para evitar agrupar cobros espurios con tolerancia amplia
                if abs(total - mov_abs) <= TOLERANCIA_COBRO:
                    intentar_match_cuotas(list(combo), ca, 'movimiento ADA')
                    encontrado = True
                    break

    # --- Estrategia 3c: cuotas sin CUIT común y sin centavos apuntando,
    # pero coincidencia EXACTA con movimiento ADA (para casos como CHAVES/COSTANZO
    # donde distintos pagadores pagan por una UF y la suma es exacta con ADA).
    # Requerimos que AL MENOS UNO de los cobros tenga centavos coincidentes con la UF,
    # para evitar falsos positivos puramente numéricos. ---
    for ca in cobros_ada:
        if ca['usado'] or ca['movimiento'] == 0:
            continue
        uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')
        mov_abs = abs(ca['movimiento'])
        disponibles = [idx for idx in range(len(cobros_banco))
                       if idx not in procesados and not cobros_banco[idx].get('es_qr')]
        if len(disponibles) < 2:
            continue
        encontrado = False
        for n in (2, 3):
            if encontrado:
                break
            if len(disponibles) < n:
                continue
            for combo in combinations(disponibles, n):
                total = sum(cobros_banco[i]['importe'] for i in combo)
                # Coincidencia EXACTA con movimiento ADA
                if abs(total - mov_abs) > TOLERANCIA_COBRO:
                    continue
                # Anchor: al menos un cobro debe tener centavos coincidentes con la UF
                if not any(obtener_centavos(cobros_banco[i]['importe']) == uf_num
                           for i in combo):
                    continue
                intentar_match_cuotas(list(combo), ca, 'movimiento ADA')
                encontrado = True
                break

    # === Estrategia 4: CENTAVOS (cobros restantes) ===
    # Para los cobros que no se resolvieron con ADA, intentar el truco de centavos
    for idx, cb in enumerate(cobros_banco):
        if idx in procesados:
            continue

        imp = imp_match(cb)
        centavos   = obtener_centavos(imp)
        cuit_banco = extraer_cuit(cb['descripcion'])

        uf_candidata = None
        estrategia   = None
        detalle_extra = ''

        # Buscar UF con mismos centavos (que no esté usada)
        # Si ADA dice que la UF no pagó (movimiento=0), solo aceptamos el match
        # cuando el importe coincide EXACTAMENTE con la deuda esperada — esto cubre
        # el caso de pagos recientes que todavía no están cargados en ADA.
        # Si el importe difiere, confiamos en ADA y no asignamos esta UF.
        for ca in cobros_ada:
            uf_num = int(re.sub(r'\D', '', ca['uf']) or '0')
            if uf_num == centavos and not ca['usado']:
                # Si ADA no tiene movimiento, exigir match exacto con la deuda
                if ca['movimiento'] == 0:
                    if abs(ca['deuda_anterior'] - imp) <= TOLERANCIA_COBRO:
                        uf_candidata = ca
                        estrategia   = 'centavos'
                    # Si no coincide exacto, confiar en ADA y no asignar
                    break
                uf_candidata = ca
                estrategia   = 'centavos'
                break

        # Fallback por importe cercano al entero
        # Solo considerar UFs que según ADA tienen algún pago registrado (movimiento != 0)
        # Si ADA dice que una UF no pagó, no la "robamos" con un match dudoso.
        if uf_candidata is None:
            imp_entero = int(imp)
            candidatos = []
            for ca in cobros_ada:
                if ca['usado']:
                    continue
                # Filtro clave: solo UFs con movimiento registrado en ADA
                if ca['movimiento'] == 0:
                    continue
                diff_entero = abs(int(ca['deuda_anterior']) - imp_entero)
                if diff_entero <= 200:
                    candidatos.append((diff_entero, ca))
            if candidatos:
                candidatos.sort(key=lambda x: x[0])
                uf_candidata = candidatos[0][1]
                estrategia   = 'importe_aproximado'

        if uf_candidata is None:
            resultados.append({
                'banco':   cb,
                'ada':     None,
                'estado':  'SIN MATCH',
                'detalle': 'No se encontró UF que corresponda a este cobro',
                'cuit':    cuit_banco,
            })
            continue

        # Determinar estado según la diferencia con la deuda
        diff_deuda = abs(uf_candidata['deuda_anterior'] - imp)
        if diff_deuda <= TOLERANCIA_COBRO:
            estado  = 'CONCILIADO'
            detalle = ''
        else:
            # Verificar si ADA muestra que la UF tiene movimientos mayores al importe
            # del banco: esto indica que hubo otros pagos anteriores y este es solo uno
            mov_abs = abs(uf_candidata['movimiento'])
            if mov_abs > 0 and mov_abs > imp and abs(mov_abs - imp) > TOLERANCIA_COBRO:
                # ADA registró más de lo que vemos en este cobro:
                # probablemente es un pago parcial dentro de un total ya registrado
                estado = 'CONCILIADO'
                detalle = (f"Pago parcial: este cobro de {imp:,.2f} es parte del total "
                           f"{mov_abs:,.2f} registrado en ADA (puede haber otros pagos previos)")
            else:
                estado = 'DIFERENCIA'
                sign   = '+' if imp > uf_candidata['deuda_anterior'] else '-'
                detalle = f"Deuda esperada: {uf_candidata['deuda_anterior']:,.2f} | Diferencia: {sign}{diff_deuda:,.2f}"
                if estrategia == 'importe_aproximado':
                    detalle += ' (match por importe, no por centavos)'

        uf_candidata['usado'] = True
        resultados.append({
            'banco':   cb,
            'ada':     uf_candidata,
            'estado':  estado,
            'detalle': detalle,
            'cuit':    cuit_banco,
        })

    # Reordenar por fila_banco
    resultados.sort(key=lambda r: r['banco']['fila_banco'])
    return resultados


def conciliar_gastos(gastos_banco, gastos_ada):
    """
    Matchea cada gasto del banco con un gasto cargado en ADA4.
    Estrategia: matching por importe con tolerancia.
    Si hay N gastos idénticos en el banco y N en ADA4, se emparejan 1-a-1
    por orden cronológico. Si las cantidades difieren, se marca AMBIGUO.
    """
    resultados = [None] * len(gastos_banco)

    # Agrupar gastos del banco por importe redondeado (para detectar duplicados exactos)
    grupos_banco = {}
    for i, gb in enumerate(gastos_banco):
        clave = round(abs(gb['importe']), 2)
        grupos_banco.setdefault(clave, []).append(i)

    # Para cada grupo, buscar candidatos en ADA
    for clave, indices in grupos_banco.items():
        # Buscar todos los gastos ADA con importe similar
        candidatos_ada = [g for g in gastos_ada
                          if not g['usado'] and abs(g['importe'] - clave) <= TOLERANCIA_GASTO]

        # Caso 1: ningún candidato → todos SIN MATCH
        if not candidatos_ada:
            for i in indices:
                gb = gastos_banco[i]
                resultados[i] = {
                    'banco':   gb,
                    'ada':     None,
                    'estado':  'SIN MATCH',
                    'detalle': 'No se encontró gasto con importe similar en ADA4',
                }
            continue

        # Caso 2: misma cantidad en ambos lados → emparejar 1-a-1 cronológicamente
        if len(indices) == len(candidatos_ada):
            # Ordenar los gastos del banco por fecha
            indices_ord = sorted(indices, key=lambda i: gastos_banco[i]['fecha'] or datetime.min)
            # Ordenar los ADA por fecha de factura (si existe) o por id
            def orden_ada(g):
                if g['fecha_fact'] and hasattr(g['fecha_fact'], 'year'):
                    return (0, g['fecha_fact'])
                return (1, g['id_gasto'])
            candidatos_ord = sorted(candidatos_ada, key=orden_ada)

            for i, ga in zip(indices_ord, candidatos_ord):
                gb = gastos_banco[i]
                ga['usado'] = True
                diff = abs(ga['importe'] - abs(gb['importe']))
                detalle = f"diff={diff:,.2f}" if diff > 0.01 else ''
                if len(indices) > 1:
                    detalle = (f"{detalle} | 1-a-1 por fecha: {len(indices)} idénticos"
                               if detalle else f"1-a-1 por fecha: {len(indices)} idénticos")
                resultados[i] = {
                    'banco':   gb,
                    'ada':     ga,
                    'estado':  'CONCILIADO',
                    'detalle': detalle,
                }
            continue

        # Caso 3: cantidades distintas (N en banco, M en ADA, N≠M) → AMBIGUO
        for i in indices:
            gb = gastos_banco[i]
            detalle = (f"{len(indices)} en banco vs {len(candidatos_ada)} en ADA4: "
                       + ', '.join([f"{c['proveedor'][:20]}({c['id_gasto']})"
                                    for c in candidatos_ada[:3]]))
            resultados[i] = {
                'banco':      gb,
                'ada':        candidatos_ada[0],
                'estado':     'AMBIGUO',
                'detalle':    detalle,
                'candidatos': candidatos_ada,
            }

    return resultados


def marcar_extracto(ruta_extracto, resultados_cobros, resultados_gastos, ruta_salida):
    """Copia el extracto original y lo marca con colores según conciliación."""
    shutil.copy2(ruta_extracto, ruta_salida)
    wb = load_workbook(ruta_salida)
    ws = wb.active

    # Combinar resultados por fila_banco
    por_fila = {}
    for r in resultados_cobros + resultados_gastos:
        por_fila[r['banco']['fila_banco']] = r

    for fila, res in por_fila.items():
        if res['estado'] == 'CONCILIADO':
            color = FILL_VERDE
        elif res['estado'] == 'PAGO EN CUOTAS':
            color = FILL_NARANJA
        elif res['estado'] in ('DIFERENCIA', 'AMBIGUO'):
            color = FILL_AMARILLO
        else:
            color = FILL_ROJO

        for col in range(1, 9):   # columnas A a H
            ws.cell(row=fila, column=col).fill = color

    wb.save(ruta_salida)


def generar_reporte(resultados_cobros, resultados_gastos, gastos_ada, cobros_ada, ruta_salida):
    """Genera un Excel con el detalle de la conciliación."""
    wb = Workbook()

    # === Hoja 1: COBROS ===
    ws = wb.active
    ws.title = "Cobros"
    headers = ['Estado', 'Fecha Banco', 'Comprobante', 'Descripción Banco',
               'Importe Banco', 'CUIT', 'UF ADA', 'Propietario ADA',
               'Deuda Esperada', 'Detalle']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal='center')

    for i, r in enumerate(resultados_cobros, 2):
        b = r['banco']
        a = r['ada']
        ws.cell(row=i, column=1, value=r['estado'])
        ws.cell(row=i, column=2, value=b['fecha']).number_format = 'DD/MM/YYYY'
        ws.cell(row=i, column=3, value=b['comprobante'])
        ws.cell(row=i, column=4, value=b['descripcion'])
        ws.cell(row=i, column=5, value=b['importe']).number_format = '#,##0.00'
        ws.cell(row=i, column=6, value=r.get('cuit', ''))
        ws.cell(row=i, column=7, value=a['uf'] if a else '')
        ws.cell(row=i, column=8, value=a['propietario'] if a else '')
        ws.cell(row=i, column=9, value=a['deuda_anterior'] if a else None).number_format = '#,##0.00'
        ws.cell(row=i, column=10, value=r['detalle'])

        # Colorear fila
        if r['estado'] == 'CONCILIADO':
            color = FILL_VERDE
        elif r['estado'] == 'PAGO EN CUOTAS':
            color = FILL_NARANJA
        elif r['estado'] == 'DIFERENCIA':
            color = FILL_AMARILLO
        else:
            color = FILL_ROJO
        for col in range(1, 11):
            ws.cell(row=i, column=col).fill = color

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 50

    # === Hoja 2: GASTOS ===
    ws2 = wb.create_sheet("Gastos")
    headers = ['Estado', 'Fecha Banco', 'Comprobante', 'Descripción Banco',
               'Importe Banco', 'Proveedor ADA', 'Importe ADA', 'Nº Factura',
               'ID ADA', 'Detalle']
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal='center')

    for i, r in enumerate(resultados_gastos, 2):
        b = r['banco']
        a = r['ada']
        ws2.cell(row=i, column=1, value=r['estado'])
        ws2.cell(row=i, column=2, value=b['fecha']).number_format = 'DD/MM/YYYY'
        ws2.cell(row=i, column=3, value=b['comprobante'])
        ws2.cell(row=i, column=4, value=b['descripcion'])
        ws2.cell(row=i, column=5, value=b['importe']).number_format = '#,##0.00'
        ws2.cell(row=i, column=6, value=a['proveedor'] if a else '')
        ws2.cell(row=i, column=7, value=a['importe'] if a else None).number_format = '#,##0.00'
        ws2.cell(row=i, column=8, value=a['nro_fact'] if a else '')
        ws2.cell(row=i, column=9, value=a['id_gasto'] if a else '')
        ws2.cell(row=i, column=10, value=r['detalle'])

        if r['estado'] == 'CONCILIADO':
            color = FILL_VERDE
        elif r['estado'] == 'AMBIGUO':
            color = FILL_AMARILLO
        else:
            color = FILL_ROJO
        for col in range(1, 11):
            ws2.cell(row=i, column=col).fill = color

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 50
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 30
    ws2.column_dimensions['G'].width = 14
    ws2.column_dimensions['H'].width = 18
    ws2.column_dimensions['I'].width = 10
    ws2.column_dimensions['J'].width = 50

    # === Hoja 3: PENDIENTES ADA4 ===
    # Lo que está en ADA4 pero no apareció en el banco (todavía no pagado/cobrado)
    ws3 = wb.create_sheet("Pendientes ADA4")
    ws3.cell(row=1, column=1, value="GASTOS CARGADOS EN ADA4 SIN MOVIMIENTO EN BANCO").font = Font(bold=True, size=12)
    headers_g = ['Proveedor', 'Importe', 'Nº Factura', 'ID', 'Descripción']
    for col, h in enumerate(headers_g, 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER

    fila = 4
    for g in gastos_ada:
        if not g['usado']:
            ws3.cell(row=fila, column=1, value=g['proveedor'])
            ws3.cell(row=fila, column=2, value=g['importe']).number_format = '#,##0.00'
            ws3.cell(row=fila, column=3, value=g['nro_fact'])
            ws3.cell(row=fila, column=4, value=g['id_gasto'])
            ws3.cell(row=fila, column=5, value=g['descripcion'][:100])
            fila += 1

    fila += 2
    ws3.cell(row=fila, column=1, value="COBROS ESPERADOS EN ADA4 SIN MOVIMIENTO EN BANCO").font = Font(bold=True, size=12)
    fila += 2
    headers_c = ['UF', 'Propietario', 'Deuda Esperada']
    for col, h in enumerate(headers_c, 1):
        c = ws3.cell(row=fila, column=col, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
    fila += 1

    for c in cobros_ada:
        if not c['usado']:
            ws3.cell(row=fila, column=1, value=c['uf'])
            ws3.cell(row=fila, column=2, value=c['propietario'])
            ws3.cell(row=fila, column=3, value=c['deuda_anterior']).number_format = '#,##0.00'
            fila += 1

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 60

    # === Hoja 4: RESUMEN ===
    ws4 = wb.create_sheet("Resumen", 0)
    ws4.cell(row=1, column=1, value="CONCILIACIÓN BANCARIA").font = Font(bold=True, size=16)
    ws4.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    tot_cob = len(resultados_cobros)
    ok_cob  = sum(1 for r in resultados_cobros if r['estado'] == 'CONCILIADO')
    cuo_cob = sum(1 for r in resultados_cobros if r['estado'] == 'PAGO EN CUOTAS')
    dif_cob = sum(1 for r in resultados_cobros if r['estado'] == 'DIFERENCIA')
    sin_cob = sum(1 for r in resultados_cobros if r['estado'] == 'SIN MATCH')

    tot_gas = len(resultados_gastos)
    ok_gas  = sum(1 for r in resultados_gastos if r['estado'] == 'CONCILIADO')
    amb_gas = sum(1 for r in resultados_gastos if r['estado'] == 'AMBIGUO')
    sin_gas = sum(1 for r in resultados_gastos if r['estado'] == 'SIN MATCH')

    ws4.cell(row=4, column=1, value="COBROS").font = Font(bold=True, size=12)
    ws4.cell(row=5, column=1, value="Total movimientos")
    ws4.cell(row=5, column=2, value=tot_cob)
    ws4.cell(row=6, column=1, value="  Conciliados")
    ws4.cell(row=6, column=2, value=ok_cob)
    ws4.cell(row=6, column=1).fill = FILL_VERDE
    ws4.cell(row=7, column=1, value="  Pago en cuotas")
    ws4.cell(row=7, column=2, value=cuo_cob)
    ws4.cell(row=7, column=1).fill = FILL_NARANJA
    ws4.cell(row=8, column=1, value="  Con diferencia")
    ws4.cell(row=8, column=2, value=dif_cob)
    ws4.cell(row=8, column=1).fill = FILL_AMARILLO
    ws4.cell(row=9, column=1, value="  Sin match")
    ws4.cell(row=9, column=2, value=sin_cob)
    ws4.cell(row=9, column=1).fill = FILL_ROJO

    ws4.cell(row=11, column=1, value="GASTOS").font = Font(bold=True, size=12)
    ws4.cell(row=12, column=1, value="Total movimientos")
    ws4.cell(row=12, column=2, value=tot_gas)
    ws4.cell(row=13, column=1, value="  Conciliados")
    ws4.cell(row=13, column=2, value=ok_gas)
    ws4.cell(row=13, column=1).fill = FILL_VERDE
    ws4.cell(row=14, column=1, value="  Ambiguos")
    ws4.cell(row=14, column=2, value=amb_gas)
    ws4.cell(row=14, column=1).fill = FILL_AMARILLO
    ws4.cell(row=15, column=1, value="  Sin match")
    ws4.cell(row=15, column=2, value=sin_gas)
    ws4.cell(row=15, column=1).fill = FILL_ROJO

    pendientes_g = sum(1 for g in gastos_ada if not g['usado'])
    pendientes_c = sum(1 for c in cobros_ada if not c['usado'])
    ws4.cell(row=17, column=1, value="PENDIENTES ADA4").font = Font(bold=True, size=12)
    ws4.cell(row=18, column=1, value="Gastos cargados sin pago bancario")
    ws4.cell(row=18, column=2, value=pendientes_g)
    ws4.cell(row=19, column=1, value="Cobros esperados sin ingreso bancario")
    ws4.cell(row=19, column=2, value=pendientes_c)

    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 10

    wb.save(ruta_salida)


def main():
    print("=" * 65)
    print("  CONCILIACIÓN BANCARIA")
    print(f"  {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}")
    print("=" * 65)

    # Verificar extracto del banco
    if not Path(EXTRACTO_BANCO).exists():
        print(f"\n❌ Archivo no encontrado (Extracto banco):\n   {EXTRACTO_BANCO}")
        input("\nPresioná Enter para cerrar...")
        sys.exit(1)

    # Buscar archivos ADA4 por prefijo
    print(f"\nBuscando archivos en: {Path(CARPETA_ADA4).resolve()}")

    INFORME_GASTOS = buscar_archivo_por_prefijo(
        CARPETA_ADA4, "InformeGastos", [".xls", ".xlsx"])
    if not INFORME_GASTOS:
        print(f"\n❌ No se encontró ningún archivo que empiece con 'InformeGastos' en {CARPETA_ADA4}")
        input("\nPresioná Enter para cerrar...")
        sys.exit(1)
    print(f"  ✓ Informe de Gastos: {Path(INFORME_GASTOS).name}")

    COBRANZA_RAPIDA = buscar_archivo_por_prefijo(
        CARPETA_ADA4, "CobranzaRapida", [".xls", ".xlsx"])
    if not COBRANZA_RAPIDA:
        print(f"\n❌ No se encontró ningún archivo que empiece con 'CobranzaRapida' en {CARPETA_ADA4}")
        input("\nPresioná Enter para cerrar...")
        sys.exit(1)
    print(f"  ✓ Cobranza Rápida:   {Path(COBRANZA_RAPIDA).name}")

    # Convertir gastos si es .xls
    ruta_gastos = INFORME_GASTOS
    if INFORME_GASTOS.lower().endswith('.xls'):
        print("\nConvirtiendo informe de gastos (.xls → .xlsx)...")
        try:
            ruta_gastos = convertir_xls(INFORME_GASTOS)
            print(f"  ✓ {Path(ruta_gastos).name}")
        except (FileNotFoundError, RuntimeError) as e:
            print(f"\n❌ {e}")
            input("\nPresioná Enter para cerrar...")
            sys.exit(1)

    # Leer los 3 archivos
    print("\nLeyendo archivos...")
    movimientos = leer_extracto(EXTRACTO_BANCO)
    gastos_ada  = leer_gastos_ada(ruta_gastos)
    cobros_ada  = leer_cobros_ada(COBRANZA_RAPIDA)

    cobros_banco = [m for m in movimientos if m['tipo'] == 'cobro']
    gastos_banco = [m for m in movimientos if m['tipo'] == 'gasto']

    print(f"  ✓ Banco:  {len(cobros_banco)} cobros, {len(gastos_banco)} gastos")
    print(f"  ✓ ADA4:   {len(cobros_ada)} cobros esperados, {len(gastos_ada)} gastos cargados")

    # Conciliar
    print("\nConciliando cobros...")
    res_cobros = conciliar_cobros(cobros_banco, cobros_ada)
    ok = sum(1 for r in res_cobros if r['estado'] in ('CONCILIADO', 'PAGO EN CUOTAS'))
    print(f"  ✓ {ok}/{len(res_cobros)} resueltos")

    print("\nConciliando gastos...")
    res_gastos = conciliar_gastos(gastos_banco, gastos_ada)
    ok = sum(1 for r in res_gastos if r['estado'] == 'CONCILIADO')
    print(f"  ✓ {ok}/{len(res_gastos)} conciliados")

    # Generar salidas
    carpeta = Path(CARPETA_SALIDA)
    carpeta.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')

    ruta_reporte = carpeta / f"ConciliacionBancaria_{timestamp}.xlsx"
    ruta_extracto_marcado = carpeta / f"Extracto_Conciliado_{timestamp}.xlsx"

    print(f"\nGenerando reporte: {ruta_reporte.name}")
    generar_reporte(res_cobros, res_gastos, gastos_ada, cobros_ada, str(ruta_reporte))

    print(f"Marcando extracto:  {ruta_extracto_marcado.name}")
    marcar_extracto(EXTRACTO_BANCO, res_cobros, res_gastos, str(ruta_extracto_marcado))

    # Resumen en consola
    print("\n" + "=" * 65)
    print("  RESUMEN")
    print("=" * 65)

    print(f"\nCOBROS ({len(res_cobros)} movimientos):")
    for r in res_cobros:
        b = r['banco']
        simbolo = {'CONCILIADO': '✓', 'PAGO EN CUOTAS': '⊞', 'DIFERENCIA': '~', 'SIN MATCH': '✗'}[r['estado']]
        uf  = r['ada']['uf'] if r['ada'] else '--'
        det = f"  ({r['detalle']})" if r['detalle'] else ''
        print(f"  {simbolo}  {b['importe']:>12,.2f}  UF {uf}  {r['estado']}{det}")

    print(f"\nGASTOS ({len(res_gastos)} movimientos):")
    for r in res_gastos:
        b = r['banco']
        simbolo = {'CONCILIADO': '✓', 'AMBIGUO': '~', 'SIN MATCH': '✗'}[r['estado']]
        prov = r['ada']['proveedor'][:30] if r['ada'] else '--'
        det  = f"  ({r['detalle']})" if r['detalle'] else ''
        print(f"  {simbolo}  {abs(b['importe']):>12,.2f}  {prov:<30}  {r['estado']}{det}")

    pend_g = sum(1 for g in gastos_ada if not g['usado'])
    pend_c = sum(1 for c in cobros_ada if not c['usado'])
    if pend_g or pend_c:
        print(f"\nPENDIENTES EN ADA4:")
        if pend_g:
            print(f"  {pend_g} gastos cargados sin pago bancario todavía")
        if pend_c:
            print(f"  {pend_c} cobros esperados sin ingresar al banco")

    print("\n" + "=" * 65)
    print("  ¡Listo! Ver archivos:")
    print(f"    1. {ruta_reporte.name}  (reporte detallado)")
    print(f"    2. {ruta_extracto_marcado.name}  (extracto con colores)")
    print("=" * 65)
    print()
    input("Presioná Enter para cerrar...")


if __name__ == "__main__":
    main()
